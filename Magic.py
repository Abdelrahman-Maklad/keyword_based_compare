# Magic Document Processing Tool
# Copyright (c) 2025 - Licensed under MIT License
"""
A tool for processing and comparing technical documents. Extracts features, keywords,
and detects changes between document versions.

Key features:
- Document text extraction
- Keyword detection
- Feature comparison
- Change tracking
- Batch processing support

Example:
    Create input.xlsx with document paths and keywords
    Configure paths in fixed_path.txt
    Run the script:
        python Magic_1.10.py

Note:
    Requires database files in ./DataBase folder:
    - keywords.xlsx: Keyword definitions
    - mapping.xlsx: Feature mapping rules
"""

import concurrent.futures
import pandas as pd
import pdfplumber
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from itertools import repeat
from tqdm import tqdm
from rich.console import Console
from rich.markdown import Markdown
from multiprocessing import freeze_support
import fitz
import re


class Word:
    """Class representing a word and its position in a document"""
    
    def __init__(self, word_text: str, word_start: float, word_end: float, 
                 word_up: float, word_down: float, word_page: int):
        """
        Initialize a word object
        
        Args:
            word_text: The text content of the word
            word_start: Starting x-coordinate
            word_end: Ending x-coordinate
            word_up: Upper y-coordinate
            word_down: Lower y-coordinate
            word_page: Page number
        """
        self.word_text = word_text
        self.word_start = word_start
        self.word_end = word_end
        self.word_up = word_up
        self.word_down = word_down
        self.word_page = word_page

    def __eq__(self, other):
        return (self.word_text == other.word_text and self.word_start == other.word_start
                and self.word_end == other.word_end and self.word_up == other.word_up)

    def __hash__(self):
        return hash((self.word_text, self.word_start, self.word_end, self.word_up))

    @property
    def word_coordinate(self):
        return self.word_up + self.word_down


class Line:
    """Class representing a line of text and its associated words"""
    
    def __init__(self, line_words: list):
        """
        Initialize a line object
        
        Args:
            line_words: List of words in the line
        """
        self.line_words = line_words

    @property
    def full_line(self):
        """Return the full line as a single string"""
        return " ".join([word.word_text for word in self.line_words])

    @property
    def line_up(self):
        """Return the uppermost y-coordinate of the line"""
        return min([word.word_up for word in self.line_words])

    @property
    def line_down(self):
        """Return the lowermost y-coordinate of the line"""
        return max([word.word_down for word in self.line_words])

    @property
    def line_start(self):
        """Return the starting x-coordinate of the line"""
        return min([word.word_start for word in self.line_words])

    @property
    def line_end(self):
        """Return the ending x-coordinate of the line"""
        return max([word.word_end for word in self.line_words])

    @property
    def line_page(self):
        """Return the page number of the line"""
        return self.line_words[0].word_page


def extract_all_words(link: str) -> list:
    """Extract all words from a document using PyMuPDF (fitz)"""
    link = link.removesuffix('\n')
    doc_1 = fitz.open(link.removesuffix('\n'))
    all_words = []
    for page_index, page in enumerate(doc_1):
        for ext_word in page.get_text("words"):
            all_words.append([ext_word, page_index])
    return all_words


def fitz_extract_all_words(link: str) -> list:
    """Extract all text from a document using PyMuPDF (fitz), page by page"""
    all_text = []
    link = link.removesuffix('\n')
    doc = fitz.open(link.removesuffix('\n'))
    for page in doc:
        for block in page.get_text("blocks"):
            line = block[4]
            new_line = re.sub(r'\s+', ' ', line.replace('\n', ' ')).strip()
            if "<image:" in new_line or new_line == '':
               continue
            all_text.append(new_line)
    return all_text

def count_of_pages(link: str) -> int:
    """Return the number of pages in a document"""
    doc = fitz.open(link)
    return doc.page_count


def extract_all_text(link: str) -> str:
    """Extract all text from a document using pdfplumber"""
    doc = pdfplumber.open(link)
    text = "\n".join([page.extract_text_simple() for page in doc.pages])
    return text.lower().strip()


def search_for_words(keywords: list, text: str) -> dict:
    """Search for keywords in the text and return matching lines"""
    matches_by_word = {keyword.strip(): [] for keyword in keywords}
    for keyword in keywords:
        matches = []
        if keyword.startswith("_"):
            matches = [text.split('\n')[int(keyword.removeprefix("_")) - 1]]
        else:
            for line in text.split("\n"):
                if keyword.strip().lower() in line.strip().lower():
                    matches.append(line.strip())

        matches = [''] if matches == [] else matches
        matches_by_word[keyword.strip()].extend(matches)
    return matches_by_word


def compare(temp_list_1: list, temp_list_2: list) -> bool:
    """Compare two lists of strings and return True if they contain the same elements"""
    match_list_1 = []
    match_list_2 = []
    for match in temp_list_1:
        match_list_1.append(match.replace(" ", ""))
    for match in temp_list_2:
        match_list_2.append(match.replace(" ", ""))
    match_list_1.sort()
    match_list_2.sort()
    if match_list_1 == match_list_2:
        return True
    elif match_list_1 != match_list_2:
        return False


def compare_and_mapping(matches_by_word1: dict, matches_by_word2: dict, keywords: list, 
                        keyword_mapping: list) -> list:
    """Compare two sets of keyword matches and map features to their corresponding database entries"""
    df_list = []
    changed_features = []
    mapped_features = []

    for keyword in keywords:

        matches_list_1 = matches_by_word1[keyword.strip()]
        matches_list_2 = matches_by_word2[keyword.strip()]

        if not compare(matches_list_1, matches_list_2):
            changed_features.append(keyword.strip())
        df_list.append("|".join(matches_by_word1[keyword.strip()]))
        df_list.append("|".join(matches_by_word2[keyword.strip()]))
        df_list.append(compare(matches_list_1, matches_list_2))

        for changed_feature in changed_features:
            for keyword_mapping_feature in keyword_mapping:
                if changed_feature.lower().strip() == keyword_mapping_feature[0].lower().strip():
                    mapped_features.append(keyword_mapping_feature[-1].lower().strip())
                    break
            else:
                mapped_features.append(changed_feature.lower().strip())

    mapped_features = list(set(mapped_features))
    mapped_features.sort()

    df_list.insert(0, ", ".join(mapped_features))
    df_list.insert(0, ", ".join(changed_features))
    df_list.insert(0, "Yes" if changed_features != [] else "No")
    return df_list


def magic_main(links: list, keywords: list, vendor_code: str, keyword_mapping: list) -> list:
    """Main processing function - extract, compare, and map keywords for a set of document links"""
    try:
        '''Get  number of pages in the document'''
        num_pages_1 = count_of_pages(links[0])
        num_pages_2 = count_of_pages(links[-1])

        if num_pages_2 > 200 or num_pages_1 > 200:
            return [links[0], links[-1], "".join(links).replace('\\\\10.199.104.50\\pdfs', ""), "Page limit exceeded"]

        text1 = "\n".join([line.lower() for line in fitz_extract_all_words(links[0])])
        text2 = "\n".join([line.lower() for line in fitz_extract_all_words(links[-1])])

        matches_by_word1 = search_for_words(keywords, text1)
        matches_by_word2 = search_for_words(keywords, text2)

        final_row = compare_and_mapping(matches_by_word1, matches_by_word2, keywords, keyword_mapping)
        final_row.insert(0, str(len(text2)))
        final_row.insert(0, str(len(text1)))
        final_row.insert(0, vendor_code)
        final_row.insert(0, "".join(links).replace('\\\\10.199.104.50\\pdfs', ""))
        final_row.insert(0, links[-1])
        final_row.insert(0, links[0])

        return final_row

    except Exception as E:
        return [links[0], links[-1], "".join(links).replace('\\\\10.199.104.50\\pdfs', ""), "Error in reading"]


def update_workbook(workbook, workbook_name, row_list):
    """Update an existing workbook with new data"""
    output_excel_to_update = workbook
    sheet = output_excel_to_update['Sheet']
    for row in row_list:
        sheet.append(row)
        for index, cell in enumerate(sheet[sheet.max_row]):
            cell.font = Font(name="Arial")
            if cell.value is not None and cell.value == False:
                cell.font = Font(name="Arial", color="FF0000")  # Red color for False
            elif cell.value is not None and cell.value == True:
                cell.font = Font(name="Arial", color="00CD00")  # Green color for True
            if index > len(row):
                break
    output_excel_to_update.save(workbook_name)


def create_workbook_and_header(workbook_name, row):
    """Create a new workbook and add a header row"""
    output_excel_created = Workbook()
    sheet = output_excel_created['Sheet']
    sheet.append(row)

    for index, cell in enumerate(sheet['1']):
        cell.font = Font(name="Arial", bold=True)
        if index > len(row):
            break
    output_excel_created.save(workbook_name)
    output_excel_created.close()  # Close the workbook before re-opening it in the next iteration
    return output_excel_created


'''----------------------------------------------------------------------------------------------------'''
if __name__ == '__main__':
    freeze_support()
    console = Console()

    title = '''# MAGIC
    > this tool is designed for the following purposes:
        1- extracting all lines with the keywords given in the keyword file "keywords.xlsx".
        2- compare between the lines and specify the changes.
        3- maps the keywords to the it's related feature in the database given in mapping file "mapping.xlsx".
    > in the "input.xlsx" input file all input are given and are grouped by the vendor code given.
    > the keywords and mapped features are grouped based on the vendor code to avoid any intersections.
    > the output is provided in separate files each named after the related supplier.
    ** each supplier must have keywords and mapping database.
    '''
    my_copyright = '''# © abdelrahman.maklad@siliconexpert.com'''
    title = Markdown(title)
    my_copyright = Markdown(my_copyright)
    console.print(title)
    console.print(my_copyright)

    title = '''▶PRESS ENTER TO START.'''
    title = Markdown(title)
    console.print(title)
    input()

    # start program
    with open("C:\D_Partition\Tools\Magic\\fixed_path.txt"
              , "r") as input_file:
        fixed_path = input_file.readline()
    fixed_path = fixed_path.strip()

    links_data_frame = pd.read_excel('C:\D_Partition\Tools\Magic\input.xlsx')
    grouped_links = links_data_frame.groupby("VENDOR_CODE")

    keywords_data_frame = pd.read_excel(f"{fixed_path}\\keywords.xlsx")
    grouped_keywords = keywords_data_frame.groupby('VENDOR_CODE')

    mapping_db = pd.read_excel(f'{fixed_path}\\mapping.xlsx')
    grouped_mapping = mapping_db.groupby('VENDOR_CODE')

    vendor_keywords_dict = {}
    vendor_links_dict = {}
    keyword_mapping_dict = {}

    for vendor_code, group in grouped_links:
        vendor_links_dict[vendor_code] = list(group[['DOCUMENT', 'LATEST']].values)

    for vendor_code, group in grouped_mapping:
        keyword_mapping_dict[vendor_code] = list(group[['KEYWORD', 'MAPPING']].values)

    for vendor_code, group in grouped_keywords:
        vendor_keywords_dict[vendor_code] = list(group['KEYWORD'])

    '''-------------------------------------------------------------------------------------------------'''
    # start program for every vendor
    for key_vendor_code in list(vendor_links_dict.keys()):
        try:
            results_list = []
            count = 1

            # create header
            header = []
            header.extend(
                ['OLD', 'LATEST', 'OLD_LATEST', 'VENDOR_CODE', 'CHAR_LEN1', 'CHAR_LEN2', 'DATA_CHANGED'
                    , 'CHANGED_FEATURES', 'MAPPED_FEATURES'])
            for word in vendor_keywords_dict[key_vendor_code]:
                header.extend(["1", "2", word.strip()])



            # create_workbook_and_header(f'{key_vendor_code}_output.xlsx', header)
            # output_excel = load_workbook(f'{key_vendor_code}_output.xlsx')

            one_time_count = 250
            links_list = vendor_links_dict[key_vendor_code]

            total_iterations = len(links_list)

            with tqdm(total=total_iterations, desc=f"Processing {key_vendor_code}".upper(), unit="row",
                      ncols=100) as progress_bar:
                count = 0
                with concurrent.futures.ProcessPoolExecutor(max_workers=5) as executor:
                    # pass links to function
                    for i in range(0, len(links_list), one_time_count):
                        batch_links = links_list[i:i + one_time_count]
                        results = executor.map(magic_main, batch_links, repeat(vendor_keywords_dict[key_vendor_code]),
                                               repeat(key_vendor_code), repeat(keyword_mapping_dict[key_vendor_code]))
                        # write results
                        for result in results:
                            results_list.append(result)
                            progress_bar.update(1)
                            count += 1

                # close workbook
                # for line in results_list:
                #     with open(f"{key_vendor_code}_output.txt", "a", encoding='utf8') as text_file:
                #             try:
                #                 text_file.write("\t".join([str(word) for word in line]) + "\n")

                #             except Exception as E:
                #                 (print(E))

                df = pandas.DataFrame(results_list, columns=header)
                df.to_excel(f'{key_vendor_code}_output.xlsx', index=False)
                    

                progress_bar.set_description(f"done {key_vendor_code}".upper())

        except Exception as E:
            print(E)
